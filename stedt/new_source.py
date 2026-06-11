#!/usr/bin/env python3
"""Onboard a new source: a contributor's wordlist file → ``data/sources/<srcabbr>/``.

The contribution unit is one folder, but its key columns can't be hand-authored: ``rn`` is a
single global sequence across every wordlist, ``lgid`` requires knowing the language table, and
a dozen workflow columns are STEDT-internal. So contributors fill a deliberately smaller
template (see TEMPLATE_COLS) and this wizard does the projection: it prompts for the
bibliography, creates this source's languages.tsv entries (a language entry is per-source —
see resolve_language), allocates fresh rns past the global maximum, and writes the folder.

The contributor file stays the source of truth until merge: re-running the wizard regenerates
the folder with freshly allocated rns, so an rn collision with a parallel merge (caught by
``stedt validate`` once the branch is up to date) is fixed by re-running, never by
hand-renumbering.

Run: ``stedt new-source --template`` for the fill-in files, ``stedt new-source FILE`` to import.
"""

import argparse
import csv
import glob
import os
import re
import subprocess
import sys
from difflib import SequenceMatcher

from stedt.paths import DATA, ROOT

csv.field_size_limit(10**7)
rel = lambda p: os.path.relpath(p, ROOT)

# The contributor-facing schema: column → (meaning, example). Everything else in wordlist.tsv is
# assigned here — rn (global sequence), lgid/language (mapped), original* (the contributor's text
# IS the original until an editor normalizes), and the workflow/analysis columns (editor work).
TEMPLATE_COLS = {
    "language": ("Language of the form. Leave the column out if the whole list is one language.", "Meithei"),
    "reflex": ("The form, exactly as the source prints it. Required.", "plet"),
    "gloss": ("Meaning, in English.", "plate"),
    "gfn": ("Grammatical function (n, v., adj, clf, …). Optional.", "n"),
    "semkey": ("STEDT thesaurus key, if known. Optional — semantic placement is editor work.", "5.11.8"),
    "page": ("Where in the source: page or entry number. Optional.", "82"),
    "note": ("Your own note on this form. Optional.", "cf. note on p.90"),
    "source_note": ("A note quoted from the source itself. Optional.", "archaic"),
}
REQUIRED = ("reflex", "gloss")


def _cols_block(width=78):
    """The TEMPLATE_COLS table as indented help-text lines, wrapped to `width`."""
    import textwrap

    w = max(map(len, TEMPLATE_COLS))
    lines = []
    for c, (doc, _) in TEMPLATE_COLS.items():
        lines.append(textwrap.fill(doc, width, initial_indent=f"  {c:<{w}}  ", subsequent_indent=" " * (w + 4)))
    return "\n".join(lines)


# The user-facing documentation. This help text is the ONE place the new-source workflow is
# documented (Luke's ruling 2026-06-11): both `stedt new-source --help` (cli.py imports it) and
# `python -m stedt.new_source --help` render it verbatim — keep every line under 78 columns.
HELP = f"""Onboard a new source: a contributor wordlist file becomes
data/sources/<srcabbr>/.

The flow: `stedt new-source --template` writes wordlist-template.tsv/.xlsx
for a contributor to fill in (Excel is fine — the .xlsx explains every
column on a second sheet). `stedt new-source <file>` then imports it:
bibliography prompts, language placement, fresh global rns, the source
folder, and finally `stedt validate` (the merge gate).

Template columns — only reflex and gloss are required:

{_cols_block()}

A `language` column is for multi-language sources: fill it on every row,
or drop the column and the wizard asks once for the whole list's language.

Language entries are per-source (a languages.tsv row is "a lect as
documented in one source"), so the import always creates this source's own
entries, placing them in the subgroup tree by same-name precedent; it
prompts only when precedent is absent or split.

The contributor file stays the source of truth. Re-running with --force
regenerates the folder with the SAME rns; if a parallel merge claimed them
first, re-run on the up-to-date branch and fresh rns are allocated — never
renumber by hand.

Every prompt has a flag, so an import can run unattended:

  stedt new-source apatani.tsv --srcabbr ABR1985 --year 1985 \\
      --author 'Abraham, P.T.' --title 'Apatani grammar' \\
      --imprint 'Mysore: CIIL' --language Apatani
"""

# Generated-file headers. These mirror validate.py's COLS — validate header-checks every file in
# data/, so any drift between the two lists fails loudly on the next run.
SOURCE_COLS = (
    "srcabbr citation author year imprint title location status dataformat format callnumber scope "
    "totalnum refonly citechk pi infascicle haveit todo proofer inputter dbprep dbload dbcheck notes"
).split()
WORDLIST_COLS = (
    "rn lgid language reflex originalreflex gloss originalgloss gfn originalgfn semkey srcid "
    "src_set_rn maintainer status analysis"
).split()
NOTES_COLS = ["rn", "type", "text"]


def die(msg):
    sys.exit(f"error: {msg}")


def read_tsv(path):
    with open(path, encoding="utf-8", newline="") as f:
        yield from csv.DictReader(f, delimiter="\t")


def write_tsv(path, header, rows):
    """The one escaping rule (data/FORMAT.md): QUOTE_MINIMAL over tab-delimited cells."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        w.writerow(header)
        w.writerows(rows)


def ask(prompt, default=""):
    tail = f" [{default}]" if default else ""
    v = input(f"{prompt}{tail}: ").strip()
    return v or default


# ───────────────────────────── the contributor file ─────────────────────────────
def _cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # Excel hands back 82 as 82.0
    return str(v).strip()


def read_contrib(path):
    """The contributor file → row dicts keyed by TEMPLATE_COLS (absent columns blank)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            die("reading .xlsx needs openpyxl (pip install openpyxl) — or save as .tsv and retry")
        ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
        raw = [[_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    elif ext in (".tsv", ".csv", ".txt"):
        # utf-8-sig: Excel's text exports lead with a BOM
        with open(path, encoding="utf-8-sig", newline="") as f:
            raw = [[_cell(c) for c in row] for row in csv.reader(f, delimiter="," if ext == ".csv" else "\t")]
    else:
        die(f"unsupported file type {ext!r} (use .tsv, .csv, or .xlsx)")
    raw = [r for r in raw if any(r)]
    if not raw:
        die(f"{path}: empty")
    header = [h.lower().replace(" ", "_") for h in raw[0]]
    unknown = [h for h in header if h not in TEMPLATE_COLS]
    if unknown:
        die(f"unknown column(s) {unknown} — the template columns are: {', '.join(TEMPLATE_COLS)}")
    missing = [c for c in REQUIRED if c not in header]
    if missing:
        die(f"missing required column(s): {', '.join(missing)}")
    rows = [dict(zip(header, r + [""] * (len(header) - len(r)))) for r in raw[1:]]
    rows = [{c: r.get(c, "") for c in TEMPLATE_COLS} for r in rows]
    blank = [i for i, r in enumerate(rows, 2) if not r["reflex"]]
    if blank:
        die(f"blank reflex on row(s) {blank[:10]} — every row needs a form")
    return rows


def write_template(outdir):
    tsv = os.path.join(outdir, "wordlist-template.tsv")
    write_tsv(tsv, list(TEMPLATE_COLS), [])
    made = [tsv]
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "wordlist"
        ws.append(list(TEMPLATE_COLS))
        # column docs + the worked example live on a second sheet, so a stray example row can
        # never ride into an import
        info = wb.create_sheet("how to fill this in")
        info.append(["column", "meaning", "example"])
        for col, (doc, example) in TEMPLATE_COLS.items():
            info.append([col, doc, example])
        info.append([])
        info.append(["Only 'reflex' and 'gloss' are required; leave anything else blank."])
        info.append(["When done: stedt new-source <this file> — or send the file to the maintainers."])
        for sheet in (ws, info):
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            for letter in "ABCDEFGH":
                sheet.column_dimensions[letter].width = 28
        info.column_dimensions["B"].width = 80
        xlsx = os.path.join(outdir, "wordlist-template.xlsx")
        wb.save(xlsx)
        made.append(xlsx)
    except ImportError:
        print("note: openpyxl not installed — skipped the .xlsx template")
    for p in made:
        print(f"wrote {p}")


# ───────────────────────────── reference tables ─────────────────────────────
def load_reference():
    langs = list(read_tsv(f"{DATA}/languages.tsv"))
    groups = {g["grpid"]: g for g in read_tsv(f"{DATA}/languagegroups.tsv")}
    semkeys = {t["semkey"] for t in read_tsv(f"{DATA}/thesaurus.tsv")}
    return langs, groups, semkeys


def next_rn(srcabbr):
    """First integer past every rn in use — the wordlists plus the orphan tables (orphan rns
    predate the digitized wordlists but still occupy the id space). The target source's own
    wordlist doesn't count: it is about to be regenerated, and skipping it makes a re-run
    reallocate the *same* rns instead of creeping past its previous self."""
    mx = 0
    for p in glob.glob(f"{DATA}/sources/*/wordlist.tsv"):
        if os.path.basename(os.path.dirname(p)) == srcabbr:
            continue
        for r in read_tsv(p):
            if r["rn"].isdigit():
                mx = max(mx, int(r["rn"]))
    for fn in ("orphan_links.tsv", "orphan_reflex_notes.tsv", "glosswords.tsv"):
        for r in read_tsv(f"{DATA}/{fn}"):
            v = r.get("rn", "")
            if v.isdigit():
                mx = max(mx, int(v))
    return mx + 1


# ───────────────────────────── language mapping ─────────────────────────────
def _lineage(groups, grpid):
    g = groups.get(grpid)
    return f"{g['grpno']} {g['name']}" if g else f"grpid {grpid}"


def _describe(l, groups):
    abbr = f" ({l['abbr']})" if l["abbr"] else ""
    return f"{l['name']}{abbr} — {_lineage(groups, l['grpid'])} — lgid {l['lgid']}"


def _candidates(name, langs):
    """Best 8 matches: exact, then substring, then fuzzy."""
    q = name.casefold()

    def score(l):
        keys = [k for k in (l["name"].casefold(), l["abbr"].casefold()) if k]
        if q in keys:
            return 3.0
        if any(q in k or k in q for k in keys):
            return 2.0
        return max((SequenceMatcher(None, q, k).ratio() for k in keys), default=0.0)

    ranked = sorted(((score(l), l) for l in langs), key=lambda t: -t[0])
    return [l for s, l in ranked if s >= 0.6][:8]


def pick_group(groups):
    while True:
        q = input("  subgroup (search by name, e.g. 'Kuki'): ").strip().casefold()
        if not q:
            continue
        hits = [g for g in groups.values() if q in g["name"].casefold() or q in g["abbr"].casefold()]
        if not hits:
            print("    no match — try another term")
            continue
        hits = sorted(hits, key=lambda g: len(g["name"]))[:12]
        for i, g in enumerate(hits, 1):
            proto = f" ({g['proto_language']})" if g["proto_language"] else ""
            print(f"    {i}. {g['grpno']} {g['name']}{proto}")
        c = input("    pick a number (or Enter to search again): ").strip()
        if c.isdigit() and 1 <= int(c) <= len(hits):
            return hits[int(c) - 1]


def create_language(name, langs, groups, srcabbr, created, grpid, abbr=""):
    lgid = str(max(int(l["lgid"]) for l in langs) + 1)
    row = {c: "" for c in langs[0].keys()}
    row.update(lgid=lgid, name=name, abbr=abbr, grpid=grpid, source=srcabbr)
    langs.append(row)
    created.append(row)
    print(f"  {name!r} → new entry (lgid {lgid}, {_lineage(groups, grpid)})")
    return row


def resolve_language(name, langs, groups, srcabbr, created, forced_grpid=None):
    """A language entry is 'a lect as documented in one source' — in the whole corpus every lgid
    is used by exactly one source, and languages.source names it. So a new source never reuses
    another source's lgid; the only open question is where its new entry sits in the subgroup
    tree. Same-named entries are the precedent: a unique precedent grpid places the entry without
    a prompt, and a re-run reuses the row this source created last time."""
    q = name.casefold()
    mine = [l for l in langs if l["source"] == srcabbr and l["name"].casefold() == q]
    if mine:
        print(f"  {name!r} → {_describe(mine[0], groups)} (this source's existing entry)")
        return mine[0]
    if forced_grpid:
        return create_language(name, langs, groups, srcabbr, created, forced_grpid)
    precedent = [l for l in langs if q in (l["name"].casefold(), l["abbr"].casefold())]
    grpids = sorted({l["grpid"] for l in precedent}, key=lambda g: -sum(1 for l in precedent if l["grpid"] == g))
    if len(grpids) == 1:
        return create_language(name, langs, groups, srcabbr, created, grpids[0])
    if grpids:
        print(f"\n  {name!r}: {len(precedent)} existing entries (one per source) sit under:")
        for i, g in enumerate(grpids, 1):
            n = sum(1 for l in precedent if l["grpid"] == g)
            print(f"    {i}. {_lineage(groups, g)} ({n} entr{'y' if n == 1 else 'ies'})")
    else:
        cands = _candidates(name, langs)
        grpids = sorted({l["grpid"] for l in cands}, key=lambda g: [l["grpid"] for l in cands].index(g))
        print(f"\n  {name!r} has no entries yet" + ("; similar names suggest:" if grpids else "."))
        for i, g in enumerate(grpids, 1):
            like = next(l for l in cands if l["grpid"] == g)
            print(f"    {i}. {_lineage(groups, g)} (like {like['name']})")
    print("    s. search the subgroup tree instead")
    while True:
        c = input("  place it under: ").strip().lower()
        if c == "s" or (c == "" and not grpids):
            grp = pick_group(groups)["grpid"]
            break
        if c.isdigit() and 1 <= int(c) <= len(grpids):
            grp = grpids[int(c) - 1]
            break
    abbr = ask(f"  abbreviation for {name!r} (optional)")
    return create_language(name, langs, groups, srcabbr, created, grp, abbr)


# ───────────────────────────── the wizard ─────────────────────────────
def bibliography(a):
    srcabbr = a.srcabbr or ask("source abbreviation (becomes the folder name, e.g. ABR1985)")
    while not re.fullmatch(r"[A-Za-z0-9._-]+", srcabbr or ""):
        srcabbr = ask("  letters, digits, . - _ only — source abbreviation")
    author = a.author if a.author is not None else ask("author(s) (e.g. 'Abraham, P.T.')")
    year = a.year if a.year is not None else ask("year")
    default_cite = f"{author.split(',')[0].strip()} {year[-2:]}" if author and len(year) >= 2 else ""
    citation = a.citation if a.citation is not None else ask("short citation", default_cite)
    title = a.title if a.title is not None else ask("title")
    imprint = a.imprint if a.imprint is not None else ask("imprint (publisher or journal)")
    return srcabbr, dict(citation=citation, author=author, year=year, title=title, imprint=imprint)


def main():
    p = argparse.ArgumentParser(description=HELP, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("wordlist", nargs="?", help="contributor wordlist (.tsv/.csv/.xlsx)")
    p.add_argument("--template", action="store_true", help="write the contributor template files and exit")
    p.add_argument("--out", default=".", help="directory for --template output")
    for flag, h in (
        ("srcabbr", "source abbreviation — becomes the folder name (e.g. ABR1985)"),
        ("citation", "short citation (e.g. 'Abraham 85'; suggested from author+year)"),
        ("author", "author(s), surname first"),
        ("year", "publication year"),
        ("title", "title of the source"),
        ("imprint", "publisher or journal"),
        ("language", "language of the whole list (only when the file has no language column)"),
    ):
        p.add_argument(f"--{flag}", help=f"{h}; prompted if omitted")
    p.add_argument("--grpid", help="subgroup id for any language entry this run creates (skips placement prompts)")
    p.add_argument("--force", action="store_true", help="regenerate an existing source folder without asking")
    p.add_argument("--no-validate", action="store_true", help="skip the validate run at the end")
    a = p.parse_args()

    if a.template:
        write_template(a.out)
        return
    if not a.wordlist:
        p.error("give a wordlist file, or --template")

    rows = read_contrib(a.wordlist)
    langs, groups, semkeys = load_reference()
    print(f"{a.wordlist}: {len(rows)} rows")
    srcabbr, bib = bibliography(a)

    outdir = f"{DATA}/sources/{srcabbr}"
    if os.path.isdir(outdir) and not a.force:
        wlp = f"{outdir}/wordlist.tsv"
        n = sum(1 for _ in read_tsv(wlp)) if os.path.exists(wlp) else 0
        if ask(f"{rel(outdir)} exists ({n} reflexes) — regenerate it from {a.wordlist}? [y/N]").lower() != "y":
            die("aborted")

    # language mapping: a 'language' column maps each distinct value; otherwise one language
    # covers the whole list
    distinct = sorted({r["language"] for r in rows if r["language"]})
    blank = sum(1 for r in rows if not r["language"])
    if distinct and blank:
        die(f"{blank} row(s) have a blank 'language' while others fill it — fill all rows or none")
    if a.grpid and a.grpid not in groups:
        die(f"grpid {a.grpid} not in languagegroups.tsv")
    created, mapping = [], {}
    if not distinct:
        name = a.language or ask("language of the whole list")
        mapping[""] = resolve_language(name, langs, groups, srcabbr, created, a.grpid)
    else:
        if a.language:
            die("the file has a 'language' column — drop --language and map interactively")
        print(f"{len(distinct)} language(s) to map:")
        for name in distinct:
            mapping[name] = resolve_language(name, langs, groups, srcabbr, created, a.grpid)

    unknown_keys = sorted({r["semkey"] for r in rows if r["semkey"] and r["semkey"] not in semkeys})
    if unknown_keys:
        print(f"warning: semkey(s) not in the thesaurus (kept; validate warns): {', '.join(unknown_keys[:8])}")
    no_gloss = sum(1 for r in rows if not r["gloss"])
    if no_gloss:
        print(f"warning: {no_gloss} row(s) have no gloss")

    first = rn = next_rn(srcabbr)
    wl, notes = [], []
    for r in rows:
        l = mapping[r["language"]] if distinct else mapping[""]
        # the contributor's text is both the working and the original reading until an editor
        # normalizes; language is derived from lgid, so write the canonical name
        wl.append(
            [str(rn), l["lgid"], l["name"], r["reflex"], r["reflex"], r["gloss"], r["gloss"],
             r["gfn"], r["gfn"], r["semkey"], r["page"], "", "", "", ""]
        )
        if r["source_note"]:
            notes.append([str(rn), "O", r["source_note"]])
        if r["note"]:
            notes.append([str(rn), "T", r["note"]])
        rn += 1

    os.makedirs(outdir, exist_ok=True)
    srow = {c: "" for c in SOURCE_COLS}
    srow.update(srcabbr=srcabbr, format="Wordlist", refonly="F", totalnum=str(len(rows)), **bib)
    write_tsv(f"{outdir}/source.tsv", SOURCE_COLS, [[srow[c] for c in SOURCE_COLS]])
    write_tsv(f"{outdir}/wordlist.tsv", WORDLIST_COLS, wl)
    if notes:
        write_tsv(f"{outdir}/notes.tsv", NOTES_COLS, notes)
    elif os.path.exists(f"{outdir}/notes.tsv"):
        os.remove(f"{outdir}/notes.tsv")  # a re-run with the note columns removed must not keep stale rns
    if created:
        write_tsv(f"{DATA}/languages.tsv", list(langs[0].keys()), [list(l.values()) for l in langs])

    noted = f", {len(notes)} notes" if notes else ""
    print(f"\nwrote {rel(outdir)}: {len(rows)} reflexes (rn {first}–{rn - 1}){noted}")
    for l in created:
        print(f"  added language {l['name']} (lgid {l['lgid']}) to languages.tsv")
    if a.no_validate:
        print("skipped validate — run `stedt validate` before opening a PR")
    else:
        print("running validate (the merge gate) …")
        code = subprocess.run([sys.executable, "-m", "stedt.validate"]).returncode
        if code != 0:
            sys.exit(code)


if __name__ == "__main__":
    main()
